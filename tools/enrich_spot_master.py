"""Enrich SPOT_OBSERVATIONS in the SPOT_INPUT master workbook.

Adds live Excel formula columns (user-approved 2026-06-11):
  - phone_lux_min / phone_lux_max / phone_lux_avg     (over cols N:R, readings 1-5)
  - lux_meter_min / lux_meter_max / lux_meter_avg     (over cols T:X, readings 1-5)
  - phone_reading_count / lux_meter_reading_count     (COUNT — makes the 3-vs-5
    protocol change explicit per row; no imputation of missing readings)
Rewrites lux_abs_diff / lux_pct_diff as formulas off the avg columns
(pct relative to the UT383 reference instrument).
Fills distance_from_window = 50/100/150 cm by triplet order (user confirmed every
session = 3 observations at increasing distance), and renames the row-2 template
example id O0001 -> O0000_EXAMPLE to remove the duplicate-id collision.

A timestamped backup copy is written next to the source before saving in place.
WINDOW_MASTER is intentionally NOT touched (user will fill its gaps).
"""

import datetime
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

MASTER = Path(
    r"C:\Users\tommy\OneDrive\Desktop\FINAL YEAR PROJECT"
    r"\LIGHT DATA COMPLETED\SPOT_INPUT_master_template.xlsx"
)

SHEET = "SPOT_OBSERVATIONS"

# Existing layout (1-indexed)
COL_OBS_ID = 1
COL_PHONE_FIRST, COL_PHONE_LAST = 14, 18      # N..R
COL_METER_FIRST, COL_METER_LAST = 20, 24      # T..X
COL_ABS_DIFF, COL_PCT_DIFF = 25, 26           # Y, Z
COL_DISTANCE = 27                              # AA

# New columns (appended after observation_note, col 31)
NEW_COLS = {
    32: "phone_lux_min",
    33: "phone_lux_max",
    34: "phone_lux_avg",
    35: "lux_meter_min",
    36: "lux_meter_max",
    37: "lux_meter_avg",
    38: "phone_reading_count",
    39: "lux_meter_reading_count",
}

DISTANCE_CYCLE = [50, 100, 150]


def reading_range(row: int, first: int, last: int) -> str:
    return f"{get_column_letter(first)}{row}:{get_column_letter(last)}{row}"


def guarded(fn: str, rng: str) -> str:
    """=IF(COUNT(rng)=0,"",FN(rng)) — blank-safe live formula."""
    return f'=IF(COUNT({rng})=0,"",{fn}({rng}))'


def main() -> int:
    if not MASTER.exists():
        print(f"ERROR: master workbook not found: {MASTER}")
        return 1

    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    backup = MASTER.with_name(f"{MASTER.stem}_BACKUP_{stamp}{MASTER.suffix}")
    shutil.copy2(MASTER, backup)
    print(f"Backup written: {backup.name}")

    wb = openpyxl.load_workbook(MASTER, data_only=False)
    ws = wb[SHEET]

    # Identify data rows: have an observation id, and at least one numeric reading.
    # Row 2 is the template example; rows with ids but no readings are pre-numbered stubs.
    data_rows = []
    example_row = None
    for r in range(2, ws.max_row + 1):
        obs = ws.cell(r, COL_OBS_ID).value
        if obs in (None, ""):
            continue
        has_reading = any(
            isinstance(ws.cell(r, c).value, (int, float))
            for c in range(COL_PHONE_FIRST, COL_PHONE_LAST + 1)
        ) or any(
            isinstance(ws.cell(r, c).value, (int, float))
            for c in range(COL_METER_FIRST, COL_METER_LAST + 1)
        )
        if r == 2:
            example_row = r
            continue
        if has_reading:
            data_rows.append(r)

    print(f"Real observation rows: {len(data_rows)}")
    if len(data_rows) % 3 != 0:
        print("ERROR: real rows are not a whole number of 3-distance triplets — aborting.")
        return 1

    # 1. Rename the row-2 template example id (kills duplicate O0001).
    if example_row is not None and ws.cell(example_row, COL_OBS_ID).value == "O0001":
        ws.cell(example_row, COL_OBS_ID).value = "O0000_EXAMPLE"
        print("Row 2 example id renamed: O0001 -> O0000_EXAMPLE")

    # 2. Headers for new columns.
    for col, name in NEW_COLS.items():
        ws.cell(1, col).value = name

    # 3. Live formulas on the example row + every real row (blank-guarded).
    formula_rows = ([example_row] if example_row else []) + data_rows
    for r in formula_rows:
        prng = reading_range(r, COL_PHONE_FIRST, COL_PHONE_LAST)
        mrng = reading_range(r, COL_METER_FIRST, COL_METER_LAST)
        ws.cell(r, 32).value = guarded("MIN", prng)
        ws.cell(r, 33).value = guarded("MAX", prng)
        ws.cell(r, 34).value = guarded("AVERAGE", prng)
        ws.cell(r, 35).value = guarded("MIN", mrng)
        ws.cell(r, 36).value = guarded("MAX", mrng)
        ws.cell(r, 37).value = guarded("AVERAGE", mrng)
        ws.cell(r, 38).value = f"=COUNT({prng})"
        ws.cell(r, 39).value = f"=COUNT({mrng})"
        # Diffs off the avg columns, relative to the UT383 reference.
        avg_p = f"{get_column_letter(34)}{r}"
        avg_m = f"{get_column_letter(37)}{r}"
        ws.cell(r, COL_ABS_DIFF).value = (
            f'=IF(OR({avg_p}="",{avg_m}=""),"",ABS({avg_p}-{avg_m}))'
        )
        ws.cell(r, COL_PCT_DIFF).value = (
            f'=IF(OR({avg_p}="",{avg_m}=""),"",ABS({avg_p}-{avg_m})/{avg_m})'
        )

    # 4. Distance fill: 50/100/150 by triplet order; never overwrite a
    #    conflicting value — verify instead.
    filled = kept = 0
    for i, r in enumerate(data_rows):
        expected = DISTANCE_CYCLE[i % 3]
        cur = ws.cell(r, COL_DISTANCE).value
        if cur in (None, ""):
            ws.cell(r, COL_DISTANCE).value = expected
            filled += 1
        elif cur == expected:
            kept += 1
        else:
            print(f"ERROR: row {r} distance {cur} != expected {expected} — aborting, no save.")
            return 1
    print(f"Distances: filled {filled}, already-correct {kept} "
          f"({len(data_rows) // 3} triplets x 50/100/150 cm)")

    try:
        wb.save(MASTER)
    except PermissionError:
        print("ERROR: workbook is open in Excel/OneDrive lock — close it and re-run.")
        return 1
    print(f"Saved: {MASTER.name}")
    print("NOTE: open once in Excel so the new formulas calculate and cache.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
