"""Batch plateau-median extraction from phyphox light streams + comparison
against the manually picked phone readings in SPOT_OBSERVATIONS.

READ-ONLY with respect to the master workbook: outputs go to
tools/analysis_output/ so the manual-vs-automated decision can be made on
evidence before anything is replaced.

Method (documented for Ch 3):
  1. Parse each phyphox .xls (sheet 'Raw Data': time s, illuminance lx).
  2. Plateau segmentation: a new segment starts when a sample deviates from the
     running segment median by > max(10%, 30 lx); segments shorter than 8
     samples (~1 s) are discarded. This automatically excludes the start/end
     tilt artifacts (phone facing the user) the field protocol identified.
  3. The observation's readings = medians of the 5 longest plateaus
     (chronological order) — robust steady-state values, one per position.
  4. File->row matching: session date parsed from the folder path, then best
     value-containment score (the manual picks were taken from the stream, so
     a row's phone values must appear in its own stream).

Outputs:
  phone_extraction_comparison.csv  per matched row: manual vs plateau readings
  phone_extraction_summary.txt     bias stats + calibration fits (manual vs
                                   plateau, same matched subset)
"""

import csv
import math
import re
import statistics
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import xlrd

ROOT = Path(r"C:\Users\tommy\OneDrive\Desktop\FINAL YEAR PROJECT\LIGHT DATA COMPLETED")
MASTER = ROOT / "SPOT_INPUT_master_template.xlsx"
OUT_DIR = Path(__file__).resolve().parent / "analysis_output"

MIN_SEG_SAMPLES = 8
REL_TOL, ABS_TOL = 0.10, 30.0
# plateaus dimmer than this fraction of the brightest plateau are treated as
# tilt artifacts (phone facing the user) and excluded — genuine positions only
# vary ~+-15% across the window width, artifacts are far darker
ARTIFACT_CUTOFF = 0.60


# --------------------------------------------------------------------------
def parse_session_date(rel: Path):
    """Top-level folder like '18-4', '27_4', '8-5', '23-5' -> (month, day)."""
    m = re.match(r"^(\d{1,2})[-_](\d{1,2})$", rel.parts[0])
    return (int(m.group(2)), int(m.group(1))) if m else None


def parse_row_date(date_cell, ts_cell):
    """(month, day) from the date col if it's a real datetime, else from the
    timestamp col (datetime or text like '15/4/2026  12:17:00 PM')."""
    for v in (date_cell, ts_cell):
        if hasattr(v, "month"):
            return (v.month, v.day)
    if isinstance(ts_cell, str):
        m = re.match(r"\s*(\d{1,2})/(\d{1,2})/(\d{4})", ts_cell)
        if m:
            return (int(m.group(2)), int(m.group(1)))
        m = re.match(r"\s*(\d{4})-(\d{1,2})-(\d{1,2})", ts_cell)
        if m:
            return (int(m.group(2)), int(m.group(3)))
    return None


def read_stream(path: Path):
    try:
        book = xlrd.open_workbook(str(path))
    except xlrd.biffh.XLRDError:
        return _read_stream_xlsx(path)  # some exports are xlsx named .xls
    sheet = None
    for s in book.sheets():
        if "raw" in s.name.lower():
            sheet = s
            break
    if sheet is None:
        sheet = book.sheet_by_index(0)
    vals = []
    for r in range(1, sheet.nrows):
        v = sheet.cell_value(r, 1)
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            vals.append(float(v))
    return vals


def _read_stream_xlsx(path: Path):
    # openpyxl validates by extension; these are xlsx bytes named .xls
    import shutil
    import tempfile
    tmp = Path(tempfile.gettempdir()) / "phyphox_tmp.xlsx"
    shutil.copy2(path, tmp)
    wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
    sheet = next((s for s in wb.worksheets if "raw" in s.title.lower()),
                 wb.worksheets[0])
    vals = []
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        v = row[0]
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            vals.append(float(v))
    wb.close()
    return vals


def plateau_medians(vals):
    """Segment into stable plateaus, drop tilt-artifact plateaus (dim relative
    to the brightest plateau), return (medians_of_5_longest_kept, n_kept,
    coverage_fraction_of_kept)."""
    segs = []
    start = 0
    for i in range(1, len(vals)):
        ref = statistics.median(vals[start:i])
        if abs(vals[i] - ref) > max(REL_TOL * ref, ABS_TOL):
            if i - start >= MIN_SEG_SAMPLES:
                segs.append((start, i))
            start = i
    if len(vals) - start >= MIN_SEG_SAMPLES:
        segs.append((start, len(vals)))
    if not segs:
        return [], 0, 0.0
    med_of = {se: statistics.median(vals[se[0]:se[1]]) for se in segs}
    brightest = max(med_of.values())
    kept = [se for se in segs if med_of[se] >= ARTIFACT_CUTOFF * brightest]
    top = sorted(sorted(kept, key=lambda se: se[1] - se[0], reverse=True)[:5])
    meds = [med_of[se] for se in top]
    coverage = sum(e - s for s, e in kept) / len(vals)
    return meds, len(kept), coverage


def ols(xs, ys):
    n = len(xs)
    sx, sy = sum(xs), sum(ys)
    sxx = sum(x * x for x in xs)
    sxy = sum(x * y for x, y in zip(xs, ys))
    slope = (n * sxy - sx * sy) / (n * sxx - sx * sx)
    intercept = (sy - slope * sx) / n
    my = sy / n
    ss_res = sum((y - (slope * x + intercept)) ** 2 for x, y in zip(xs, ys))
    ss_tot = sum((y - my) ** 2 for y in ys)
    r = (n * sxy - sx * sy) / math.sqrt(
        (n * sxx - sx * sx) * (n * sum(y * y for y in ys) - sy * sy))
    return slope, intercept, 1 - ss_res / ss_tot, r


# --------------------------------------------------------------------------
def main() -> int:
    OUT_DIR.mkdir(exist_ok=True)

    # --- master rows ------------------------------------------------------
    wb = openpyxl.load_workbook(MASTER, data_only=False)
    ws = wb["SPOT_OBSERVATIONS"]
    rows = []
    for r in range(3, ws.max_row + 1):
        obs = ws.cell(r, 1).value
        if obs in (None, ""):
            continue
        phone = [v for c in range(14, 19)
                 if isinstance((v := ws.cell(r, c).value), (int, float))]
        meter = [v for c in range(20, 25)
                 if isinstance((v := ws.cell(r, c).value), (int, float))]
        if not phone:
            continue
        md = parse_row_date(ws.cell(r, 5).value, ws.cell(r, 4).value)
        rows.append({
            "obs": obs, "win": ws.cell(r, 3).value,
            "month": md[0] if md else None,
            "day": md[1] if md else None,
            "dist": ws.cell(r, 27).value,
            "phone": phone, "meter": meter,
        })
    by_date = defaultdict(list)
    undated = []
    for x in rows:
        if x["month"] is None:
            undated.append(x)
        else:
            by_date[(x["month"], x["day"])].append(x)
    print(f"master rows with phone readings: {len(rows)} "
          f"(dated {len(rows) - len(undated)}, undated {len(undated)})")

    # --- streams ----------------------------------------------------------
    files = [p for p in ROOT.rglob("*.xls") if "light" in p.name.lower()]
    print(f"phyphox files: {len(files)}")

    results, unmatched_files, bad_files = [], [], []
    claimed = {}
    for path in sorted(files):
        rel = path.relative_to(ROOT)
        date = parse_session_date(rel)
        try:
            vals = read_stream(path)
        except Exception as e:  # corrupt/odd file — record and move on
            bad_files.append((str(rel), str(e)))
            continue
        if len(vals) < 30:
            bad_files.append((str(rel), f"only {len(vals)} samples"))
            continue
        vset = set(vals)
        # dated rows from the same session date + all undated rows; fall back
        # to every row when the folder date can't be parsed
        candidates = (by_date.get(date, []) + undated) if date else rows
        best, best_score, second = None, -1, -1
        for x in candidates:
            score = sum(1 for p in x["phone"] if float(p) in vset)
            if score > best_score:
                best, best_score, second = x, score, best_score
            elif score > second:
                second = score
        if best is None or best_score < max(2, len(best["phone"]) - 1):
            unmatched_files.append((str(rel), best_score))
            continue
        meds, nseg, coverage = plateau_medians(vals)
        results.append({
            "obs": best["obs"], "win": best["win"], "dist": best["dist"],
            "file": str(rel), "match_score": f"{best_score}/{len(best['phone'])}",
            "ambiguous": second == best_score,
            "manual": best["phone"], "meter": best["meter"],
            "plateau": meds, "n_segments": nseg,
            "coverage": round(coverage, 2), "n_samples": len(vals),
        })
        claimed.setdefault(best["obs"], []).append(str(rel))

    dup_rows = {k: v for k, v in claimed.items() if len(v) > 1}
    matched_obs = set(claimed)
    unmatched_rows = [x["obs"] for x in rows if x["obs"] not in matched_obs]

    # --- comparison CSV ----------------------------------------------------
    comp = OUT_DIR / "phone_extraction_comparison.csv"
    with comp.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["obs_id", "window", "distance_cm", "file", "match_score",
                    "ambiguous", "n_samples", "n_segments", "coverage",
                    "manual_readings", "plateau_readings",
                    "manual_avg", "plateau_avg", "manual_vs_plateau_pct"])
        for x in sorted(results, key=lambda y: y["obs"]):
            ma = statistics.mean(x["manual"])
            pa = statistics.mean(x["plateau"]) if x["plateau"] else None
            w.writerow([
                x["obs"], x["win"], x["dist"], x["file"], x["match_score"],
                x["ambiguous"], x["n_samples"], x["n_segments"], x["coverage"],
                ";".join(str(round(v)) for v in x["manual"]),
                ";".join(str(round(v)) for v in x["plateau"]),
                round(ma, 1), round(pa, 1) if pa else "",
                round((ma - pa) / pa * 100, 1) if pa else "",
            ])

    # --- summary + calibrations on the matched subset ----------------------
    ok = [x for x in results if x["plateau"] and x["meter"] and not x["ambiguous"]]
    diffs = [(statistics.mean(x["manual"]) - statistics.mean(x["plateau"]))
             / statistics.mean(x["plateau"]) * 100 for x in ok]
    M = [statistics.mean(x["meter"]) for x in ok]
    man = [statistics.mean(x["manual"]) for x in ok]
    plat = [statistics.mean(x["plateau"]) for x in ok]
    s1, i1, r2_1, r1 = ols(man, M)
    s2, i2, r2_2, r2 = ols(plat, M)

    lines = [
        f"files: {len(files)}  parsed-ok: {len(files) - len(bad_files)}  "
        f"matched: {len(results)}  unmatched files: {len(unmatched_files)}  "
        f"bad files: {len(bad_files)}",
        f"master rows matched: {len(matched_obs)}/{len(rows)}  "
        f"(unmatched rows: {len(unmatched_rows)})",
        f"rows where >1 file claimed the same obs: {len(dup_rows)}",
        "",
        f"manual avg vs plateau avg (n={len(ok)} clean matches):",
        f"  median bias: {statistics.median(diffs):+.1f}%   "
        f"IQR: [{statistics.quantiles(diffs, n=4)[0]:+.1f}%, "
        f"{statistics.quantiles(diffs, n=4)[2]:+.1f}%]",
        "",
        "calibration vs UT383 on the SAME matched subset:",
        f"  manual picks : meter = {s1:.4f} * phone + {i1:.1f}   "
        f"R2={r2_1:.4f}  r={r1:.4f}",
        f"  plateau meds : meter = {s2:.4f} * phone + {i2:.1f}   "
        f"R2={r2_2:.4f}  r={r2:.4f}",
        "",
        "unmatched rows: " + ", ".join(unmatched_rows[:30]),
        "unmatched files:",
    ] + [f"  {p} (best score {s})" for p, s in unmatched_files[:30]] + [
        "bad files:",
    ] + [f"  {p}: {e}" for p, e in bad_files[:30]]

    (OUT_DIR / "phone_extraction_summary.txt").write_text(
        "\n".join(lines), encoding="utf-8")
    print("\n".join(lines[:14]))
    print(f"\nWrote {comp.name} and phone_extraction_summary.txt")
    return 0


if __name__ == "__main__":
    sys.exit(main())
