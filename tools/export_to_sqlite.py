#!/usr/bin/env python3
"""
export_to_sqlite.py  --  Build step for the Spot-Based Indoor Plant Recommender.

Reads the FROZEN Excel source of truth (PLANT_MASTER + RAW_EVIDENCES + LOOKUPS) and
emits a bundled, read-only runtime database:

    android/app/src/main/assets/plant_db.sqlite

The build is GATED on data-integrity checks (the same ones used while curating the
dataset): wrong row counts, orphan foreign keys, missing evidence references, LOOKUP
violations, or empty source URLs all abort the export with a non-zero exit code so a
broken dataset can never ship.

Run:  python tools/export_to_sqlite.py
Deps: openpyxl  (pip install openpyxl)
"""

import os
import sys
import sqlite3
import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    sys.exit("ERROR: openpyxl not installed. Run: pip install openpyxl")

# --------------------------------------------------------------------------- paths
FYP = r"C:\Users\tommy\OneDrive\Desktop\FINAL YEAR PROJECT"
PM_XLSX = os.path.join(FYP, r"(LATEST 6-7-final-v3) PLANT_MASTER_updated_with_zhang2023_support.xlsx")
RE_XLSX = os.path.join(FYP, r"(LATEST 6-7-final-v3) RAW_EVIDENCES_updated_with_zhang2023_support.xlsx")
LK_XLSX = os.path.join(FYP, r"PLANT DATA\LOOKUPS.xlsx")

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT_DIR = os.path.join(REPO, "android", "app", "src", "main", "assets")
OUT_DB = os.path.join(OUT_DIR, "plant_db.sqlite")

SCHEMA_VERSION = 1
EXPECTED_PLANTS = 31
EXPECTED_EVIDENCE = 171

# --------------------------------------------------------------------------- schema
# PLANT_MASTER: header row 1, data row 2+  (27 columns, 1-indexed)
PLANT_COLS = [
    "plant_id", "scientific_name_accepted", "common_name_main", "family",
    "indoor_relevance", "tropical_relevance",
    "maintenance_lux_min", "maintenance_lux_max", "preferred_lux_min", "preferred_lux_max",
    "maintenance_ppfd_min", "maintenance_ppfd_max", "preferred_ppfd_min", "preferred_ppfd_max",
    "dli_min", "dli_max", "photoperiod_min", "photoperiod_max",
    "primary_evidence_ids", "supporting_evidence_ids",
    "final_confidence", "value_status", "inheritance_status", "threshold_selection_note",
    "shade_category", "aspect_orientation", "direct_sun_tolerance",
]
PLANT_NUMERIC = {
    "maintenance_lux_min", "maintenance_lux_max", "preferred_lux_min", "preferred_lux_max",
    "maintenance_ppfd_min", "maintenance_ppfd_max", "preferred_ppfd_min", "preferred_ppfd_max",
    "dli_min", "dli_max", "photoperiod_min", "photoperiod_max",
}

# RAW_EVIDENCES: field names in row 2 (cols 1-29) + row 1 (cols 30-31), data row 3+
EVIDENCE_COLS = [
    "evidence_id", "plant_key", "scientific_name_raw", "scientific_name_accepted",
    "common_raw_name", "source_title", "source_url", "source_type", "source_section",
    "date_accessed", "specificity_level", "context_type", "raw_light_text",
    "raw_value_min", "raw_value_max", "raw_unit", "sunlight_hours_min", "sunlight_hours_max",
    "shade_category", "photoperiod_min", "photoperiod_max", "DLI_min", "DLI_max",
    "aspect_orientation", "evidence_grade", "traceability_status", "used_in_final_threshold",
    "reason_if_not_used", "notes", "direct_sun_tolerance", "direct_sun_tolerance_note",
]
EVIDENCE_NUMERIC = {
    "raw_value_min", "raw_value_max", "sunlight_hours_min", "sunlight_hours_max",
    "photoperiod_min", "photoperiod_max", "DLI_min", "DLI_max",
}

# LOOKUPS: table-name in row 1 at these columns; codes from row 3 down
LOOKUP_TABLE_COLS = {
    "LOOKUP_SOURCE_TYPE": 1, "LOOKUP_CONTEXT_TYPE": 6, "LOOKUP_SPECIFICITY_LEVEL": 10,
    "LOOKUP_EVIDENCE_GRADE": 13, "LOOKUP_USED_IN_FINAL_THRESHOLD": 18,
    "LOOKUP_SHADE_CATEGORY": 20, "LOOKUP_ASPECT_ORIENTATION": 24,
    "LOOKUP_EXPOSURE_CONDITION": 26, "LOOKUP_UNITS": 29, "LOOKUP_PROXY_STATUS": 33,
    "LOOKUP_CONFIDENCE_LEVEL": 36, "LOOKUP_SOURCE_PRIORITY": 38,
    "LOOKUP_DIRECT_SUN_TOLERANCE": 42,
}


def num(v):
    """Coerce to float/int if the cell is numeric, else None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return None


def txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s != "" else None


def split_codes(v):
    """Multi-value cells are ';'-joined. Return list of trimmed tokens."""
    if v is None:
        return []
    return [t.strip() for t in str(v).split(";") if t.strip()]


# --------------------------------------------------------------------------- read
def read_plants():
    ws = openpyxl.load_workbook(PM_XLSX, read_only=True, data_only=True).active
    rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value in (None, ""):
            continue
        rec = {}
        for i, col in enumerate(PLANT_COLS, start=1):
            v = ws.cell(r, i).value
            rec[col] = num(v) if col in PLANT_NUMERIC else txt(v)
        rows.append(rec)
    return rows


def read_evidence():
    ws = openpyxl.load_workbook(RE_XLSX, read_only=True, data_only=True).active
    rows = []
    for r in range(3, ws.max_row + 1):
        if ws.cell(r, 1).value in (None, ""):
            continue
        rec = {}
        for i, col in enumerate(EVIDENCE_COLS, start=1):
            v = ws.cell(r, i).value
            rec[col] = num(v) if col in EVIDENCE_NUMERIC else txt(v)
        rows.append(rec)
    return rows


def read_lookups():
    ws = openpyxl.load_workbook(LK_XLSX, read_only=True, data_only=True).active
    out = []  # (category, code, sort_order)
    for name, col in LOOKUP_TABLE_COLS.items():
        order = 0
        for r in range(3, ws.max_row + 1):
            code = ws.cell(r, col).value
            if code is None or str(code).strip() == "":
                break
            out.append((name, str(code).strip(), order))
            order += 1
    return out


# --------------------------------------------------------------------------- checks
def integrity_gate(plants, evidence, lookups):
    errors = []
    by_cat = {}
    for cat, code, _ in lookups:
        by_cat.setdefault(cat, set()).add(code)

    # row counts
    if len(plants) != EXPECTED_PLANTS:
        errors.append(f"plant count {len(plants)} != expected {EXPECTED_PLANTS}")
    if len(evidence) != EXPECTED_EVIDENCE:
        errors.append(f"evidence count {len(evidence)} != expected {EXPECTED_EVIDENCE}")

    plant_ids = {p["plant_id"] for p in plants}
    evidence_ids = {e["evidence_id"] for e in evidence}

    # duplicate ids
    if len(plant_ids) != len(plants):
        errors.append("duplicate plant_id present")
    if len(evidence_ids) != len(evidence):
        errors.append("duplicate evidence_id present")

    # orphan evidence.plant_key
    for e in evidence:
        if e["plant_key"] not in plant_ids:
            errors.append(f"orphan evidence plant_key {e['evidence_id']} -> {e['plant_key']}")

    # PLANT_MASTER evidence refs must exist in evidence
    for p in plants:
        for col in ("primary_evidence_ids", "supporting_evidence_ids"):
            for ref in split_codes(p[col]):
                if ref not in evidence_ids:
                    errors.append(f"{p['plant_id']}.{col} -> missing evidence {ref}")

    # empty source_url
    for e in evidence:
        if not e["source_url"]:
            errors.append(f"evidence {e['evidence_id']} has empty source_url")

    # LOOKUP compliance -- plant
    def chk(val, cat, who):
        if val is None:
            return
        valid = by_cat.get(cat, set())
        for tok in split_codes(val):
            if tok not in valid:
                errors.append(f"{who}: '{tok}' not in {cat}")

    for p in plants:
        chk(p["final_confidence"], "LOOKUP_CONFIDENCE_LEVEL", f"plant {p['plant_id']}.final_confidence")
        chk(p["shade_category"], "LOOKUP_SHADE_CATEGORY", f"plant {p['plant_id']}.shade_category")
        chk(p["aspect_orientation"], "LOOKUP_ASPECT_ORIENTATION", f"plant {p['plant_id']}.aspect_orientation")
        chk(p["direct_sun_tolerance"], "LOOKUP_DIRECT_SUN_TOLERANCE", f"plant {p['plant_id']}.direct_sun_tolerance")

    for e in evidence:
        w = f"evidence {e['evidence_id']}"
        chk(e["source_type"], "LOOKUP_SOURCE_TYPE", w + ".source_type")
        chk(e["specificity_level"], "LOOKUP_SPECIFICITY_LEVEL", w + ".specificity_level")
        chk(e["context_type"], "LOOKUP_CONTEXT_TYPE", w + ".context_type")
        chk(e["evidence_grade"], "LOOKUP_EVIDENCE_GRADE", w + ".evidence_grade")
        chk(e["used_in_final_threshold"], "LOOKUP_USED_IN_FINAL_THRESHOLD", w + ".used_in_final_threshold")
        chk(e["raw_unit"], "LOOKUP_UNITS", w + ".raw_unit")
        chk(e["shade_category"], "LOOKUP_SHADE_CATEGORY", w + ".shade_category")
        chk(e["aspect_orientation"], "LOOKUP_ASPECT_ORIENTATION", w + ".aspect_orientation")
        chk(e["direct_sun_tolerance"], "LOOKUP_DIRECT_SUN_TOLERANCE", w + ".direct_sun_tolerance")

    return errors


# --------------------------------------------------------------------------- write
def build_db(plants, evidence, lookups):
    os.makedirs(OUT_DIR, exist_ok=True)
    if os.path.exists(OUT_DB):
        os.remove(OUT_DB)
    con = sqlite3.connect(OUT_DB)
    cur = con.cursor()

    def coltype(c, numeric):
        return "REAL" if c in numeric else "TEXT"

    cur.execute("CREATE TABLE meta (key TEXT PRIMARY KEY, value TEXT)")
    cur.execute(
        "CREATE TABLE plant (%s, PRIMARY KEY(plant_id))"
        % ", ".join(f"{c} {coltype(c, PLANT_NUMERIC)}" for c in PLANT_COLS)
    )
    cur.execute(
        "CREATE TABLE evidence (%s, PRIMARY KEY(evidence_id), "
        "FOREIGN KEY(plant_key) REFERENCES plant(plant_id))"
        % ", ".join(f"{c} {coltype(c, EVIDENCE_NUMERIC)}" for c in EVIDENCE_COLS)
    )
    cur.execute("CREATE TABLE lookup (category TEXT, code TEXT, sort_order INTEGER, "
                "PRIMARY KEY(category, code))")

    cur.executemany(
        "INSERT INTO plant VALUES (%s)" % ", ".join("?" * len(PLANT_COLS)),
        [[p[c] for c in PLANT_COLS] for p in plants],
    )
    cur.executemany(
        "INSERT INTO evidence VALUES (%s)" % ", ".join("?" * len(EVIDENCE_COLS)),
        [[e[c] for c in EVIDENCE_COLS] for e in evidence],
    )
    cur.executemany("INSERT INTO lookup VALUES (?,?,?)", lookups)

    cur.execute("CREATE INDEX idx_evidence_plant ON evidence(plant_key)")

    meta = {
        "schema_version": str(SCHEMA_VERSION),
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "plant_rows": str(len(plants)),
        "evidence_rows": str(len(evidence)),
        "lookup_rows": str(len(lookups)),
        "source_plant_master": os.path.basename(PM_XLSX),
        "source_raw_evidences": os.path.basename(RE_XLSX),
        "source_lookups": os.path.relpath(LK_XLSX, FYP),
    }
    cur.executemany("INSERT INTO meta VALUES (?,?)", list(meta.items()))

    con.commit()
    con.close()
    return meta


def main():
    for path in (PM_XLSX, RE_XLSX, LK_XLSX):
        if not os.path.exists(path):
            sys.exit(f"ERROR: source not found: {path}")

    print("Reading Excel source of truth ...")
    plants = read_plants()
    evidence = read_evidence()
    lookups = read_lookups()
    print(f"  plant={len(plants)}  evidence={len(evidence)}  lookup_codes={len(lookups)}")

    print("Running integrity gate ...")
    errors = integrity_gate(plants, evidence, lookups)
    if errors:
        print(f"  FAILED ({len(errors)} issue(s)) -- DB NOT written:")
        for e in errors[:40]:
            print("   -", e)
        sys.exit(1)
    print("  PASS: counts, orphans, refs, LOOKUP codes, URLs all clean.")

    print("Building SQLite ...")
    meta = build_db(plants, evidence, lookups)
    size_kb = os.path.getsize(OUT_DB) / 1024
    print(f"  wrote {OUT_DB}  ({size_kb:.0f} KB)")
    print("  meta:", {k: meta[k] for k in ("schema_version", "plant_rows", "evidence_rows", "lookup_rows")})
    print("DONE.")


if __name__ == "__main__":
    main()
