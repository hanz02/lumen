"""Blank the 11 mean-imputed lux-meter cells in SPOT_OBSERVATIONS.

Each cell was detected as equal (within rounding) to the mean of the row's other
four readings while those four were well spread — the user confirmed these were
mean-filled where the UT383 display photo was too blurred to transcribe.
Treatment per the no-imputation policy: discard the value (rows become honest
4-reading rows; blank-guarded MIN/MAX/AVERAGE/COUNT formulas adapt) and append
a dated audit note to observation_note. Borderline rows O0039/O0092 are NOT
touched (possibly genuine readings; user to verify against photos).

Waits for the workbook to be closed in Excel (up to 10 min), then backs up,
verifies every expected cell value, edits, and saves. Aborts on any mismatch.
"""

import datetime
import shutil
import sys
import time
from pathlib import Path

import openpyxl

MASTER = Path(
    r"C:\Users\tommy\OneDrive\Desktop\FINAL YEAR PROJECT"
    r"\LIGHT DATA COMPLETED\SPOT_INPUT_master_template.xlsx"
)
SHEET = "SPOT_OBSERVATIONS"
COL_NOTE = 31
WAIT_SECONDS = 600

# (row, col, expected_value, reading_number) — col 20..24 = lux_meter readings 1..5
CELLS = [
    (95, 20, 459, 1),    # O0086
    (126, 20, 288, 1),   # O0114
    (131, 20, 352, 1),   # O0119
    (133, 24, 508, 5),   # O0121
    (135, 20, 185, 1),   # O0123 (two cells flagged in this row)
    (135, 22, 185, 3),   # O0123
    (137, 22, 88, 3),    # O0125
    (148, 24, 1710, 5),  # O0134
    (163, 20, 2111, 1),  # O0147
    (211, 21, 246, 2),   # O0194
    (221, 24, 63, 5),    # O0203
]


def wait_for_unlock(path: Path, timeout_s: int) -> bool:
    deadline = time.time() + timeout_s
    while time.time() < deadline:
        try:
            with path.open("r+b"):
                return True
        except PermissionError:
            time.sleep(15)
    return False


def main() -> int:
    if not MASTER.exists():
        print(f"ERROR: not found: {MASTER}")
        return 1
    print("Waiting for the workbook to be closed in Excel...")
    if not wait_for_unlock(MASTER, WAIT_SECONDS):
        print("ERROR: workbook still locked after 10 min — re-run after closing Excel.")
        return 1

    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    backup = MASTER.with_name(f"{MASTER.stem}_BACKUP_{stamp}{MASTER.suffix}")
    shutil.copy2(MASTER, backup)
    print(f"Backup written: {backup.name}")

    wb = openpyxl.load_workbook(MASTER, data_only=False)
    ws = wb[SHEET]

    # Verify every cell before touching anything.
    for r, c, expected, _n in CELLS:
        actual = ws.cell(r, c).value
        if actual != expected:
            print(f"ERROR: row {r} col {c}: expected {expected}, found {actual!r} — "
                  f"aborting, nothing changed.")
            return 1

    today = datetime.date.today().isoformat()
    by_row: dict[int, list[tuple[int, int, int]]] = {}
    for r, c, v, n in CELLS:
        by_row.setdefault(r, []).append((c, v, n))

    for r, items in by_row.items():
        obs = ws.cell(r, 1).value
        for c, _v, _n in items:
            ws.cell(r, c).value = None
        nums = " & ".join(f"#{n} ({v} lx)" for _c, v, n in items)
        note = (f"{today}: lux_meter reading {nums} discarded — display photo "
                f"unreadable; value had been mean-filled. Genuine readings retained.")
        cur = ws.cell(r, COL_NOTE).value
        ws.cell(r, COL_NOTE).value = f"{cur} | {note}" if cur not in (None, "") else note
        print(f"  {obs} (row {r}): blanked {nums}")

    try:
        wb.save(MASTER)
    except PermissionError:
        print("ERROR: file re-locked during edit — close Excel and re-run.")
        return 1
    print(f"Saved. {len(CELLS)} cells blanked across {len(by_row)} observations.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
