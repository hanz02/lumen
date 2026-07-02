"""Phone-vs-UT383 agreement analysis over SPOT_OBSERVATIONS (Ch 4 data source).

Reads the field master workbook and computes everything from the RAW reading
cells (never from formula cells — openpyxl has no cached values for formulas
written by enrich_spot_master.py until Excel recalculates).

Outputs (tools/analysis_output/):
  agreement_summary.csv   overall + by sky / lux band / distance / reading-count:
                          n, MAE, MAPE, median pct bias, Pearson r, OLS fit
                          (slope, intercept, R^2), Bland-Altman (bias, SD, LoA)
  distance_falloff.csv    per session: meter avg at 50/100/150 cm + falloff ratios
  within_spot_spread.csv  per row: reading counts, min/max range, range % of avg
                          (Ch 4 spread statistic should filter reading_count=5)
  calibration_constants.txt  the constants for src/engine/config.ts, with the
                          sensitivity fit (5-reading rows only) alongside

Method citations (Ch 3/4): Bland & Altman 1986 for agreement; OLS calibration
precedent in smartphone-lux literature (see plan section 8).
"""

import csv
import datetime
import math
import statistics
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

MASTER = Path(
    r"C:\Users\tommy\OneDrive\Desktop\FINAL YEAR PROJECT"
    r"\LIGHT DATA COMPLETED\SPOT_INPUT_master_template.xlsx"
)
OUT_DIR = Path(__file__).resolve().parent / "analysis_output"

SHEET = "SPOT_OBSERVATIONS"
COL_OBS_ID, COL_WIN, COL_DATE, COL_TIME = 1, 3, 5, 6
COL_SKY = 7
COL_PHONE_FIRST, COL_PHONE_LAST = 14, 18
COL_METER_FIRST, COL_METER_LAST = 20, 24
COL_DISTANCE = 27

DISTANCE_CYCLE = [50, 100, 150]


def nums(ws, r, first, last):
    return [v for c in range(first, last + 1)
            if isinstance((v := ws.cell(r, c).value), (int, float))]


def pearson(xs, ys):
    mx, my = statistics.mean(xs), statistics.mean(ys)
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    den = math.sqrt(sum((x - mx) ** 2 for x in xs) * sum((y - my) ** 2 for y in ys))
    return num / den


def ols(xs, ys):
    """meter = slope * phone + intercept; returns (slope, intercept, r2)."""
    n = len(xs)
    sx, sy = sum(xs), sum(ys)
    sxx = sum(x * x for x in xs)
    sxy = sum(x * y for x, y in zip(xs, ys))
    slope = (n * sxy - sx * sy) / (n * sxx - sx * sx)
    intercept = (sy - slope * sx) / n
    my = sy / n
    ss_res = sum((y - (slope * x + intercept)) ** 2 for x, y in zip(xs, ys))
    ss_tot = sum((y - my) ** 2 for y in ys)
    return slope, intercept, 1 - ss_res / ss_tot


def agreement_stats(rows):
    """rows: list of dicts with p_avg/m_avg. Returns a stats dict."""
    P = [x["p_avg"] for x in rows]
    M = [x["m_avg"] for x in rows]
    diffs = [p - m for p, m in zip(P, M)]            # phone - meter
    pcts = [(p - m) / m * 100 for p, m in zip(P, M)]
    slope, intercept, r2 = ols(P, M)
    bias = statistics.mean(diffs)
    sd = statistics.stdev(diffs) if len(diffs) > 1 else 0.0
    return {
        "n": len(rows),
        "mae_lux": round(statistics.mean(abs(d) for d in diffs), 1),
        "mape_pct": round(statistics.mean(abs(p) for p in pcts), 2),
        "median_pct_bias": round(statistics.median(pcts), 2),
        "pearson_r": round(pearson(P, M), 4),
        "fit_slope": round(slope, 4),
        "fit_intercept": round(intercept, 1),
        "fit_r2": round(r2, 4),
        "ba_bias_lux": round(bias, 1),
        "ba_sd_lux": round(sd, 1),
        "ba_loa_low": round(bias - 1.96 * sd, 1),
        "ba_loa_high": round(bias + 1.96 * sd, 1),
    }


def kfold_session_cv(rows, k=5, min_phone=0):
    """Session-level k-fold cross-validation of the phone->meter calibration.

    Holds out WHOLE sessions (consecutive triplets) so a session's 50/100/150 cm
    rows never split across train/test (no leakage). For each fold: refit OLS on
    the training sessions, predict the held-out rows, and collect the held-out
    errors. This validates the calibration generalises beyond its fitting set;
    the SHIPPED constants stay fit on ALL rows (computed separately in main()).

    `min_phone` restricts fit + scoring to rows with phone >= that value, so the
    validation can be reported over the range where calibration is actually
    applied (the app returns raw lux below LUX_CALIBRATION.validMinLux). Session
    grouping is preserved regardless (no leakage). Returns (per_fold, overall).
    """
    sessions = [rows[i:i + 3] for i in range(0, len(rows), 3)]
    k = min(k, len(sessions))
    if k < 2:
        return [], {"folds": 0, "n_test_total": 0,
                    "cv_mae": None, "cv_rmse": None, "cv_mape": None}
    folds = [[] for _ in range(k)]
    for idx, sess in enumerate(sessions):
        folds[idx % k].append(sess)            # deterministic round-robin

    per_fold = []
    abs_all, sq_all, pct_all = [], [], []
    for fi in range(k):
        test_rows = [r for s in folds[fi] for r in s if r["p_avg"] >= min_phone]
        train_rows = [r for j in range(k) if j != fi
                      for s in folds[j] for r in s if r["p_avg"] >= min_phone]
        if len(train_rows) < 2 or not test_rows:
            continue
        slope, intercept, _ = ols([r["p_avg"] for r in train_rows],
                                  [r["m_avg"] for r in train_rows])
        a, q, p = [], [], []
        for r in test_rows:
            err = (slope * r["p_avg"] + intercept) - r["m_avg"]
            a.append(abs(err)); q.append(err * err)
            p.append(abs(err) / r["m_avg"] * 100)
        abs_all += a; sq_all += q; pct_all += p
        per_fold.append({
            "fold": fi + 1, "n_train": len(train_rows), "n_test": len(test_rows),
            "slope": round(slope, 4), "intercept": round(intercept, 1),
            "mae": round(statistics.mean(a), 1),
            "rmse": round(math.sqrt(statistics.mean(q)), 1),
            "mape": round(statistics.mean(p), 2),
        })
    overall = {
        "folds": len(per_fold), "n_test_total": len(abs_all),
        "cv_mae": round(statistics.mean(abs_all), 1) if abs_all else None,
        "cv_rmse": round(math.sqrt(statistics.mean(sq_all)), 1) if sq_all else None,
        "cv_mape": round(statistics.mean(pct_all), 2) if pct_all else None,
    }
    return per_fold, overall


def main() -> int:
    if not MASTER.exists():
        print(f"ERROR: master workbook not found: {MASTER}")
        return 1
    OUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(MASTER, data_only=True, read_only=True)
    ws = wb[SHEET]
    # read_only sheets need explicit dimensions reset for full iteration
    ws.reset_dimensions() if hasattr(ws, "reset_dimensions") else None

    rows = []
    grid = [list(r) for r in ws.iter_rows(min_row=2, max_col=31, values_only=True)]
    for i, vals in enumerate(grid):
        r = i + 2
        obs = vals[COL_OBS_ID - 1]
        if obs in (None, "") or r == 2:   # skip blank + the row-2 template example
            continue
        phone = [v for v in vals[COL_PHONE_FIRST - 1:COL_PHONE_LAST]
                 if isinstance(v, (int, float))]
        meter = [v for v in vals[COL_METER_FIRST - 1:COL_METER_LAST]
                 if isinstance(v, (int, float))]
        if not phone or not meter:
            continue                      # pre-numbered stub rows
        rows.append({
            "row": r, "obs": obs, "win": vals[COL_WIN - 1],
            "date": vals[COL_DATE - 1], "time": vals[COL_TIME - 1],
            "sky": vals[COL_SKY - 1] or "unknown",
            "dist": vals[COL_DISTANCE - 1],
            "p_n": len(phone), "m_n": len(meter),
            "p_avg": statistics.mean(phone), "m_avg": statistics.mean(meter),
            "p_min": min(phone), "p_max": max(phone),
            "m_min": min(meter), "m_max": max(meter),
        })
    print(f"Paired observations: {len(rows)}")
    if len(rows) % 3:
        print("WARNING: row count is not a whole number of triplets.")

    # Distance: use the sheet value if present, else derive from triplet order
    # (enrichment may not have been saved yet if the workbook was locked).
    for i, x in enumerate(rows):
        if x["dist"] in (None, ""):
            x["dist"] = DISTANCE_CYCLE[i % 3]

    # --- agreement_summary.csv ---------------------------------------------
    def band(m):
        return "<500" if m < 500 else "500-2000" if m < 2000 else ">=2000"

    groups = [("overall", "all", rows)]
    for key, label_fn in (("sky", lambda x: x["sky"]),
                          ("lux_band", lambda x: band(x["m_avg"])),
                          ("distance_cm", lambda x: str(x["dist"])),
                          ("reading_count", lambda x: f"{min(x['p_n'], x['m_n'])}")):
        sub = defaultdict(list)
        for x in rows:
            sub[label_fn(x)].append(x)
        for label in sorted(sub):
            groups.append((key, label, sub[label]))

    summary_path = OUT_DIR / "agreement_summary.csv"
    with summary_path.open("w", newline="", encoding="utf-8") as f:
        w = None
        for key, label, sub in groups:
            if len(sub) < 3:
                continue
            stats = {"split": key, "group": label, **agreement_stats(sub)}
            if w is None:
                w = csv.DictWriter(f, fieldnames=list(stats))
                w.writeheader()
            w.writerow(stats)
    print(f"Wrote {summary_path.name}")

    # --- distance_falloff.csv ----------------------------------------------
    falloff_path = OUT_DIR / "distance_falloff.csv"
    with falloff_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["session", "window", "date", "time", "sky",
                    "meter_avg_50cm", "meter_avg_100cm", "meter_avg_150cm",
                    "ratio_50_to_100", "ratio_50_to_150",
                    "phone_avg_50cm", "phone_avg_100cm", "phone_avg_150cm"])
        ratios = []
        for i in range(0, len(rows) - 2, 3):
            a, b, c = rows[i], rows[i + 1], rows[i + 2]
            r53 = a["m_avg"] / c["m_avg"] if c["m_avg"] else None
            if r53:
                ratios.append(r53)
            w.writerow([i // 3 + 1, a["win"], a["date"], a["time"], a["sky"],
                        round(a["m_avg"], 1), round(b["m_avg"], 1), round(c["m_avg"], 1),
                        round(a["m_avg"] / b["m_avg"], 2) if b["m_avg"] else "",
                        round(r53, 2) if r53 else "",
                        round(a["p_avg"], 1), round(b["p_avg"], 1), round(c["p_avg"], 1)])
    print(f"Wrote {falloff_path.name} "
          f"(median 50->150 falloff: {statistics.median(ratios):.2f}x)")

    # --- within_spot_spread.csv --------------------------------------------
    spread_path = OUT_DIR / "within_spot_spread.csv"
    with spread_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["obs_id", "window", "distance_cm", "phone_n", "meter_n",
                    "phone_range_lux", "phone_range_pct_of_avg",
                    "meter_range_lux", "meter_range_pct_of_avg"])
        for x in rows:
            w.writerow([x["obs"], x["win"], x["dist"], x["p_n"], x["m_n"],
                        round(x["p_max"] - x["p_min"], 1),
                        round((x["p_max"] - x["p_min"]) / x["p_avg"] * 100, 1),
                        round(x["m_max"] - x["m_min"], 1),
                        round((x["m_max"] - x["m_min"]) / x["m_avg"] * 100, 1)])
    full5 = [x for x in rows if x["p_n"] >= 5 and x["m_n"] >= 5]
    sp = statistics.median((x["p_max"] - x["p_min"]) / x["p_avg"] * 100 for x in full5)
    sm = statistics.median((x["m_max"] - x["m_min"]) / x["m_avg"] * 100 for x in full5)
    print(f"Wrote {spread_path.name} (5-reading rows n={len(full5)}: "
          f"median range phone {sp:.0f}%, meter {sm:.0f}% of avg)")

    # --- calibration constants + sensitivity --------------------------------
    all_fit = agreement_stats(rows)
    sens_rows = [x for x in rows if x["p_n"] >= 5 and x["m_n"] >= 5]
    sens_fit = agreement_stats(sens_rows)
    VALID_MIN_LUX = 200  # mirrors src/engine/config.ts LUX_CALIBRATION.validMinLux
    per_fold, cv = kfold_session_cv(rows, k=5, min_phone=VALID_MIN_LUX)
    _, cv_all = kfold_session_cv(rows, k=5, min_phone=0)
    cal_path = OUT_DIR / "calibration_constants.txt"
    today = datetime.date.today().isoformat()
    cal_path.write_text(
        f"Generated {today} by analyze_spot_observations.py\n"
        f"Calibration: meter_lux = slope * phone_lux + intercept\n\n"
        f"ALL rows (n={all_fit['n']}):\n"
        f"  slope={all_fit['fit_slope']}  intercept={all_fit['fit_intercept']}"
        f"  R2={all_fit['fit_r2']}  r={all_fit['pearson_r']}\n\n"
        f"SENSITIVITY — 5-reading rows only (n={sens_fit['n']}):\n"
        f"  slope={sens_fit['fit_slope']}  intercept={sens_fit['fit_intercept']}"
        f"  R2={sens_fit['fit_r2']}  r={sens_fit['pearson_r']}\n\n"
        f"SESSION {cv['folds']}-FOLD CROSS-VALIDATION (held-out, refit per fold):\n"
        f"  In validated range (phone>={VALID_MIN_LUX} lx, n_test={cv['n_test_total']}):\n"
        f"    MAE={cv['cv_mae']} lux  RMSE={cv['cv_rmse']} lux  MAPE={cv['cv_mape']}%\n"
        f"  All rows incl. sub-{VALID_MIN_LUX} lx (n_test={cv_all['n_test_total']}):\n"
        f"    MAE={cv_all['cv_mae']} lux  RMSE={cv_all['cv_rmse']} lux  MAPE={cv_all['cv_mape']}%\n"
        f"  (MAPE is inflated below the valid range, where a small lux error is a\n"
        f"  large %, which is exactly why the app returns raw lux there.)\n"
        f"  Validates the calibration generalises beyond its fitting set. The\n"
        f"  SHIPPED constants stay the ALL-rows fit above; this is a separate,\n"
        f"  reported validation number (see calibration_crossval.csv per fold).\n\n"
        f"Use the ALL-rows constants in src/engine/config.ts LUX_CALIBRATION.\n"
        f"Device-specific (Samsung S21+), valid ~200-6000 lx indoor daylight.\n",
        encoding="utf-8",
    )
    print(f"Wrote {cal_path.name}")
    print(f"  ALL  : meter = {all_fit['fit_slope']} * phone + "
          f"{all_fit['fit_intercept']}  (R2={all_fit['fit_r2']}, n={all_fit['n']})")
    print(f"  5-rdg: meter = {sens_fit['fit_slope']} * phone + "
          f"{sens_fit['fit_intercept']}  (R2={sens_fit['fit_r2']}, n={sens_fit['n']})")

    # --- calibration_crossval.csv (session k-fold held-out validation) ------
    cv_path = OUT_DIR / "calibration_crossval.csv"
    with cv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["fold", "n_train", "n_test", "fold_slope", "fold_intercept",
                    "mae_lux", "rmse_lux", "mape_pct"])
        for r in per_fold:
            w.writerow([r["fold"], r["n_train"], r["n_test"], r["slope"],
                        r["intercept"], r["mae"], r["rmse"], r["mape"]])
        w.writerow([])
        w.writerow([f"HELD-OUT (phone>={VALID_MIN_LUX} lx)", "", cv["n_test_total"],
                    "", "", cv["cv_mae"], cv["cv_rmse"], cv["cv_mape"]])
        w.writerow(["HELD-OUT (all rows)", "", cv_all["n_test_total"], "", "",
                    cv_all["cv_mae"], cv_all["cv_rmse"], cv_all["cv_mape"]])
    print(f"Wrote {cv_path.name} (session {cv['folds']}-fold CV, in-range >="
          f"{VALID_MIN_LUX} lx: held-out MAE={cv['cv_mae']} lux, "
          f"MAPE={cv['cv_mape']}%, RMSE={cv['cv_rmse']} lux, n_test={cv['n_test_total']})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
