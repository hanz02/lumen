"""Second imputation-removal pass: blank the 10 lux_meter reading#5 cells that
contain a live ROUNDUP(mean-of-other-4) table formula — the user's original
mean-fill for observations whose 5th display photo was unreadable. Complements
remove_imputed_meter_readings.py (which handled the typed-value imputations).
Same policy: discard, never impute; rows become honest 4-reading rows."""

import datetime
import shutil
import sys
from pathlib import Path

import openpyxl

MASTER = Path(
    r"C:\Users\tommy\OneDrive\Desktop\FINAL YEAR PROJECT"
    r"\LIGHT DATA COMPLETED\SPOT_INPUT_master_template.xlsx"
)
SHEET = "SPOT_OBSERVATIONS"
COL_NOTE = 31
COL_METER5 = 24

ROWS = [164, 165, 170, 184, 185, 186, 190, 195, 207, 217]


def main() -> int:
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    backup = MASTER.with_name(f"{MASTER.stem}_BACKUP_{stamp}{MASTER.suffix}")
    shutil.copy2(MASTER, backup)
    print(f"Backup written: {backup.name}")

    wb = openpyxl.load_workbook(MASTER, data_only=False)
    ws = wb[SHEET]

    for r in ROWS:
        v = ws.cell(r, COL_METER5).value
        if not (isinstance(v, str) and v.startswith("=ROUNDUP(")):
            print(f"ERROR: row {r} reading#5 is {v!r}, not the expected ROUNDUP "
                  f"formula — aborting, nothing changed.")
            return 1

    today = datetime.date.today().isoformat()
    for r in ROWS:
        ws.cell(r, COL_METER5).value = None
        note = (f"{today}: lux_meter reading #5 discarded — display photo "
                f"unreadable; cell held a ROUNDUP mean-of-other-4 formula "
                f"(mean-fill). 4 genuine readings retained.")
        cur = ws.cell(r, COL_NOTE).value
        ws.cell(r, COL_NOTE).value = f"{cur} | {note}" if cur not in (None, "") else note
        print(f"  {ws.cell(r, 1).value} (row {r}): formula cell blanked")

    try:
        wb.save(MASTER)
    except PermissionError:
        print("ERROR: workbook open in Excel — close it and re-run.")
        return 1
    print(f"Saved. {len(ROWS)} formula cells blanked.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
